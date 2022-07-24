from typing import Dict, Tuple, List
import sqlite3, os, shutil, sys, tempfile, time
import pandas as pd
import numpy as np
from uuid import uuid4
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

import browser

Row = Tuple[Cell]
Observation = Tuple[str, float]
Series = List[Observation]
Definition = Tuple[str, str, str, str]


def scrape(dir = None) -> list[pd.DataFrame]:
  """Fetch time series data from RBNZ website.

  Returns:
      list[pd.DataFrame]: DataFrames containing series definitions and time 
      series data.
  """
  if not dir:
    dir, _ = download()
  defs, data = _process_dir(dir)
  shutil.rmtree(dir)
  exclude = ['EXRT.YS45.ZZB17']
  df1 = _defs_to_pd(defs).astype(object).replace(np.nan, None).drop_duplicates(subset = ['id'])
  df2 = _data_to_pd(data).astype(object).replace(np.nan, None).drop_duplicates(subset = ['id', 'date'])
  return [df1.loc[df1['id'].isin(exclude) == False], df2.loc[df2['id'].isin(exclude) == False]]


def download() -> Tuple[str, list[str]]:
  """Download all Excel spreadsheets and save to random folder.

  Returns:
    A tuple containing the download folder name, and the list of all downloaded
    files.
  """
  wd = f"{tempfile.gettempdir()}/{str(uuid4())}"
  os.mkdir(wd)
  driver = browser.getHeadlessDriver(wd)
  driver.get("https://www.rbnz.govt.nz/statistics/series/data-file-index-page")
  rows = driver.find_elements(By.CSS_SELECTOR, ".table-wrapper tbody > tr")
  links = list(map(_process_row, rows))
  
  for xs in links:
    for x in xs:
      driver.get(x)
      # to be complient with RBNZ terms of use:
      time.sleep(60) 
  
  files = []
  for xs in links:
    for x in xs:
      f = f"{wd}/{_filename_from_url(x)}"
      if os.path.exists(f):
        files.append(f)
  
  driver.close()
  return (wd, files)


def _process_row(row: Row) -> list[str]:
  links = row.find_element(By.CSS_SELECTOR, ":nth-child(4)").find_elements(By.TAG_NAME, "a")
  urls = list(map(lambda x: x.get_attribute("href"), links))
  return list(filter(lambda x: "xlsx" in x.lower(), urls))


def _filename_from_url(url: str) -> str:
  tmp = url.split("/")[-1]
  return tmp.split("?")[0]


def import_series_data(filename: str, res: dict[str, Series] = {}) -> dict[str, Series]:
  """Import time series data from Excel file.

  Args:
      filename (str): Excel file.
      res (dict[str, Series], optional): Dictionary of time series. Defaults to {}.

  Returns:
      dict[str, Series]: (Updated) dictionary of time series.
  """
  try:
    wb = load_workbook(filename = filename)
  except:
    return res
  
  if not ('Data' in wb.sheetnames and 'Series Definitions' in wb.sheetnames):
    return res
  
  ws = wb['Data']
  id_row = next(x for i,x in enumerate(ws.rows) if i==4)
  ids = [x.value.upper() for i,x in enumerate(id_row) if i > 0 and x.value]  
  dates = [str(x[0].value) for i, x in enumerate(ws.rows) if i > 4 and x[0].internal_value]
  
  def safe_value(cell):
    if y.value:
      if str(y.value).strip() == "-":
        return None
      elif str(y.value).strip() == "":
        return None
      else:
        return y.value
    else: 
      return None
  
  for i,x in enumerate(ws.rows):
    if i >= 5 and (i < len(dates) + 5):
      for j,y in enumerate(x):
        if j > 0 and j <= len(ids):
          if ids[j - 1] in res:
            res[ids[j - 1]].append((dates[i - 5], safe_value(y)))
          else:
            res[ids[j - 1]] = [(dates[i - 5], safe_value(y))]
  
  return res


def import_series_definitions(filename: str, res: dict[str, Definition] = {}) -> dict[str, Definition]:
  """Import series definitions from Excel file.

  Args:
      filename (str): Excel file
      res (dict[str, Definition], optional): Dictionary of series definitions. Defaults to {}.

  Returns:
      dict[str, Definition]: (Updated) dictionary of series definitions.
  """
  try:
    wb = load_workbook(filename = filename)
  except:
    return res
  if not ('Data' in wb.sheetnames and 'Series Definitions' in wb.sheetnames):
    return res
  ws = wb['Series Definitions']
  for i,row in enumerate(ws.rows):
    if i > 0:
      def f(x, upper = False): 
        if not x.value:
          return None
        if upper:
          return x.value.strip().upper()
        else:
          return x.value.strip() 
      id = f(row[2], True)
      group = f(row[0])
      series = f(row[1])
      unit = f(row[3], True)
      note = f(row[4])
      res[id] = (group, series, unit, note)
  return res


def _process_dir(dir: str) -> Tuple[dict[str, Definition], dict[str, Series]]:
  data = {}
  defs = {}
  for f in os.listdir(dir):
    if "xlsx" in f:
      import_series_data(f"{dir}/{f}", data)
      import_series_definitions(f"{dir}/{f}", defs)
  return (defs, data)


def _defs_to_pd(x: dict[str, Definition]) -> pd.DataFrame:
  def f(k,v):
    return pd.DataFrame({
      'group': [v[0]],
      'id': [k],
      'name': [v[1]],
      'unit': [v[2]],
      'note': [v[3]]
    })
  return pd.concat([f(k,v) for k,v in x.items()])


def _data_to_pd(x: dict[str, Series]) -> pd.DataFrame:
  def f(k,v):
    d = sorted(v.copy(), key = lambda x: x[0])
    return pd.DataFrame({
      'id': [k] * len(v),
      'date': [x[0] for x in v],
      'value': [x[1] for x in v]
    })
  return pd.concat([f(k,v) for k,v in x.items()])


def connect(path: str = None) -> sqlite3.Connection:
  """Get database connection.
  
  Args:
    path: Path to sqlite database.
    
  Returns:
    Database connection of class `pyodbc.Connection`.
  """
  conn = sqlite3.connect(path if path else ":memory:")
  return conn


def df_to_sql(conn: sqlite3.Connection, df: pd.core.frame.DataFrame, table: str) -> bool:
  """Save DataFrame to SQLite db.
  
  Args:
    conn: Database connection.
    df: DataFrame. 
    table: Table name.
    
  Returns:
    True if table exists, False if not.
  """
  return df.to_sql(table, conn, if_exists = 'replace', index = False) > 0


def main():
  defs, series = scrape()
  conn = connect(sys.argv[1])
  df_to_sql(conn, defs, "series_definition")
  df_to_sql(conn, series, "series")
  conn.execute("create index idx1 on definitions(id);")
  conn.execute("create index idx2 on series(id, date);")
  conn.commit()
  conn.close()


if __name__ == "__main__":
    main()